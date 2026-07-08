import sys
import re
import json
import datetime as dt
from pathlib import Path
import os

from openpyxl import load_workbook
"""
labs_parser.py

קורא קובץ Excel של לוחות מעבדות ומייבא ל-Firestore:
- מזהה טבלאות לפי כותרות
- משייך מעבדות לקורסים (קוד + שם)
- שומר מפגשים: תאריך, יום, שעה, קבוצה, מרצה
- תומך במספר גיליונות ובמבנים משתנים
- תאים ממוזגים + fill-down ל-staff/date בתוך אותה טבלת קורס

שימוש:
python labs_parser.py <file_path> <year_id> <year_label> <semester>
python labs_parser.py <file_path> --dry-run   (ניתוח בלבד, ללא כתיבה ל-Firestore)
"""

# ==============================
# Required headers (UNCHANGED)
# ==============================
REQUIRED_HEADERS = {
    "staff": ["שם המרצה", "מרצה"],
    "group": ["קבוצת מעבדה", "קבוצה"],
    "time": ["שעה"],
    "day": ["יום"],
    "date": ["תאריך"],
    "sessionNo": ["מס' מע'", "מספר מעבדה", "מס׳ מע", "מס' מעבדה"],
    "sessionName": ["שם המקצוע", "שם הקורס"],
}

# ==============================
# Helpers
# ==============================
def norm(x):
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()


def cell_value(ws, row, col):
    """Return cell value, resolving merged ranges to the top-left anchor."""
    if not col or col < 1:
        return None
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return cell.value


def build_header(ws, row, max_col):
    cells = [norm(cell_value(ws, row, c)) for c in range(1, max_col + 1)]
    mapping = {}
    hits = 0

    for key, variants in REQUIRED_HEADERS.items():
        for i, txt in enumerate(cells, start=1):
            if txt and any(v in txt for v in variants):
                mapping[key] = i
                hits += 1
                break

    return mapping, hits


def is_table_header_row(hits):
    return hits >= 5


def extract_course_from_line(text):
    t = norm(text)
    if not t:
        return None, None

    m = re.search(r"(.+?)\s*[-–]\s*(\d{4,6})$", t)
    if m:
        return m.group(2), norm(m.group(1))

    m = re.search(r"^(\d{4,6})\s*[-–]\s*(.+)$", t)
    if m:
        return m.group(1), norm(m.group(2))

    return None, None


def find_course_title_near(ws, header_row, max_col):
    for r in range(header_row - 1, max(1, header_row - 8), -1):
        line = " ".join(
            norm(cell_value(ws, r, c))
            for c in range(1, max_col + 1)
            if cell_value(ws, r, c)
        )
        code, name = extract_course_from_line(line)
        if code and name:
            return code, name

    return None, None


def _iso(d, m, y):
    y = str(y)
    if len(y) == 2:
        y = "20" + y
    return "%s-%02d-%02d" % (y, int(m), int(d))


def normalize_date_cell(value):
    """
    Normalize a raw date cell to (date, dateEnd) as ISO yyyy-mm-dd strings.
    Handles datetime cells, d.m.yy text, day ranges (14-16.6.26) and
    cross-month ranges (14.6-2.7.26). Unrecognized text is returned verbatim
    with an empty dateEnd so the quality report can flag it.
    """
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d"), ""

    t = norm(value)
    if not t:
        return "", ""
    t = re.sub(r"^[א-ת]'?\s*", "", t).strip()

    m = re.match(r"^(\d{4}-\d{2}-\d{2})", t)
    if m:
        return m.group(1), ""

    m = re.match(r"^(\d{1,2})[./](\d{1,2})[./](\d{2,4})$", t)
    if m:
        return _iso(m.group(1), m.group(2), m.group(3)), ""

    m = re.match(r"^(\d{1,2})\s*[-–]\s*(\d{1,2})[./](\d{1,2})[./](\d{2,4})$", t)
    if m:
        return (
            _iso(m.group(1), m.group(3), m.group(4)),
            _iso(m.group(2), m.group(3), m.group(4)),
        )

    m = re.match(r"^(\d{1,2})[./](\d{1,2})\s*[-–]\s*(\d{1,2})[./](\d{1,2})[./](\d{2,4})$", t)
    if m:
        return (
            _iso(m.group(1), m.group(2), m.group(5)),
            _iso(m.group(3), m.group(4), m.group(5)),
        )

    return t, ""


def normalize_time_cell(value):
    """Normalize a time cell to HH:MM; text values pass through as-is."""
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M")
    if isinstance(value, dt.time):
        return "%02d:%02d" % (value.hour, value.minute)
    return norm(value)


def _looks_like_date(text):
    t = norm(text)
    if not t:
        return False
    if re.search(r"\d{1,2}[./]\d{1,2}", t):
        return True
    if re.search(r"\d{4}-\d{2}-\d{2}", t):
        return True
    return False


def _is_fillable_staff(text):
    t = norm(text)
    if not t:
        return False
    header_words = ("שם המרצה", "מרצה", "תאריך", "יום", "שעה", "קבוצה")
    if t in header_words:
        return False
    return True


# ==============================
# Core parsing (no Firebase)
# ==============================
def parse_workbook_data(path: Path) -> dict:
    """
    Parse lab workbook into courses_map.
    Structure matches Firestore payload: { courseCode: { courseCode, courseName, labs[] } }
    """
    wb = load_workbook(path, data_only=True)
    courses_map = {}

    for ws in wb.worksheets:
        r = 1
        while r <= ws.max_row:
            header, hits = build_header(ws, r, ws.max_column)

            if not is_table_header_row(hits):
                r += 1
                continue

            course_code, course_name = find_course_title_near(ws, r, ws.max_column)
            if not course_code:
                r += 1
                continue

            courses_map.setdefault(
                course_code,
                {
                    "courseCode": str(course_code),
                    "courseName": course_name,
                    "labs": [],
                },
            )

            last_date = ""
            last_date_end = ""
            last_staff = ""

            rr = r + 1
            while rr <= ws.max_row:
                raw_date = cell_value(ws, rr, header.get("date"))
                date_val, date_end_val = normalize_date_cell(raw_date)
                day_val = norm(cell_value(ws, rr, header.get("day")))

                session_val = ""
                if header.get("sessionNo"):
                    session_val = norm(cell_value(ws, rr, header["sessionNo"]))

                if not session_val and header.get("sessionName"):
                    session_val = norm(cell_value(ws, rr, header["sessionName"]))

                if not session_val and not date_val and not day_val:
                    break

                staff_val = ""
                if header.get("staff"):
                    staff_val = norm(cell_value(ws, rr, header["staff"]))

                # fill-down within the same course table only
                if not date_val and last_date and _looks_like_date(last_date):
                    date_val = last_date
                    date_end_val = last_date_end
                if not staff_val and last_staff and _is_fillable_staff(last_staff):
                    staff_val = last_staff

                if _looks_like_date(date_val):
                    last_date = date_val
                    last_date_end = date_end_val
                if _is_fillable_staff(staff_val):
                    last_staff = staff_val

                lab = {
                    "session": session_val,
                    "date": date_val,
                    "day": day_val,
                    "group": norm(cell_value(ws, rr, header.get("group"))),
                    "time": normalize_time_cell(cell_value(ws, rr, header.get("time"))),
                    "staff": [staff_val] if staff_val else [],
                }
                if date_end_val:
                    lab["dateEnd"] = date_end_val

                courses_map[course_code]["labs"].append(lab)
                rr += 1

            r = rr

    return courses_map


def count_lab_quality_issues(courses_map: dict) -> dict:
    """Count lab rows missing fields (same rules as knowledge-check)."""
    total = 0
    missing_date = 0
    missing_staff = 0
    missing_any = 0
    unparsed_dates = 0
    rows_with_issues = []

    for code, course in courses_map.items():
        for lab in course.get("labs") or []:
            total += 1
            missing = []
            if not lab.get("date"):
                missing.append("date")
                missing_date += 1
            elif not re.match(r"^\d{4}-\d{2}-\d{2}$", lab["date"]):
                missing.append("dateFormat")
                unparsed_dates += 1
            if not lab.get("day"):
                missing.append("day")
            if not lab.get("time"):
                missing.append("time")
            if lab.get("group") in (None, ""):
                missing.append("group")
            staff = lab.get("staff") or []
            if not staff or (isinstance(staff, list) and not len(staff)):
                missing.append("staff")
                missing_staff += 1
            if missing:
                missing_any += 1
                rows_with_issues.append({
                    "courseCode": code,
                    "courseName": course.get("courseName"),
                    "session": lab.get("session"),
                    "group": lab.get("group"),
                    "missing": missing,
                })

    return {
        "totalRecords": total,
        "rowsWithMissingFields": missing_any,
        "missingDate": missing_date,
        "missingStaff": missing_staff,
        "unparsedDates": unparsed_dates,
        "rowsWithIssues": rows_with_issues,
    }


def build_report(courses_map: dict) -> dict:
    return {
        "ok": True,
        "totalCourses": len(courses_map),
        "totalLabs": sum(len(c.get("labs") or []) for c in courses_map.values()),
        "courses": [
            {
                "courseCode": c["courseCode"],
                "courseName": c["courseName"],
                "labCount": len(c.get("labs") or []),
            }
            for c in courses_map.values()
        ],
        "quality": count_lab_quality_issues(courses_map),
    }


# ==============================
# Firebase write
# ==============================
def _init_firebase():
    # Imported lazily so --dry-run works without firebase-admin installed
    import firebase_admin
    from firebase_admin import credentials, firestore

    if firebase_admin._apps:
        return firestore.client()
    cred = credentials.Certificate({
        "type": "service_account",
        "project_id": os.environ["FIREBASE_PROJECT_ID"],
        "client_email": os.environ["FIREBASE_CLIENT_EMAIL"],
        "private_key": os.environ["FIREBASE_PRIVATE_KEY"].replace("\\n", "\n"),
        "token_uri": "https://oauth2.googleapis.com/token",
    })
    firebase_admin.initialize_app(cred)
    return firestore.client()


def parse_workbook(path: Path, year_id: str, year_label: str, semester: str):
    from firebase_admin import firestore

    db = _init_firebase()
    courses_map = parse_workbook_data(path)

    year_ref = db.collection("lab_schedule").document(year_id)
    year_ref.set(
        {
            "year": year_label,
            "updatedAt": firestore.SERVER_TIMESTAMP,
        },
        merge=True,
    )

    semester_ref = year_ref.collection("semesters").document(str(semester))
    semester_ref.set(
        {
            "semester": int(semester),
            "updatedAt": firestore.SERVER_TIMESTAMP,
            "courses": courses_map,
        },
        merge=False,
    )

    return courses_map


# ==============================
# ENTRY POINT
# ==============================
if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    dry_run = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--dry-run"]

    if dry_run:
        if len(args) < 1:
            raise Exception("Usage: labs_parser.py <file_path> --dry-run")
        courses_map = parse_workbook_data(Path(args[0]))
    else:
        if len(args) < 4:
            raise Exception("Usage: labs_parser.py <file_path> <year_id> <year_label> <semester>")
        courses_map = parse_workbook(Path(args[0]), args[1], args[2], args[3])

    # Last stdout line is the JSON payload consumed by uploadAdmin.js
    print(json.dumps({"report": build_report(courses_map), "courses": courses_map}))
